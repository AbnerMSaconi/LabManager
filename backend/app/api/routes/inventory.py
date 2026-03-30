import json
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional

from ..deps import get_db, RoleChecker, get_current_user
from .sse import broadcast
from ...models.base_models import (
    ItemModel, Reservation, ReservationItem, ReservationStatus,
    User, UserRole, InstitutionLoan, InventoryMovement, AuditLog
)
from ...schemas.reservation_schemas import ItemModelCreate, ItemModelUpdate

router = APIRouter(prefix="/api/v1/inventory", tags=["inventário"])

# Papéis que podem apenas visualisar e operar o estoque do dia a dia (Checkout, Empréstimos, etc)
_ALMOX = [UserRole.DTI_TECNICO, UserRole.DTI_ESTAGIARIO, UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]

# Papéis que podem CRIAR, EDITAR ou RESOLVER MANUTENÇÕES do estoque
_MANAGE_ITEMS = [UserRole.DTI_TECNICO, UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]


@router.get("/models")
async def list_item_models(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(ItemModel).filter(ItemModel.deleted_at == None).all()


@router.get("/models/available")
async def list_available_models(
    date: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import datetime as _dt
    try:
        target_date = _dt.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Data inválida. Use YYYY-MM-DD.")

    active_statuses = [
        ReservationStatus.PENDENTE.value, ReservationStatus.APROVADO.value,
        ReservationStatus.AGUARDANDO_SOFTWARE.value, ReservationStatus.EM_USO.value,
        ReservationStatus.APROVADO_COM_RESSALVAS.value,
    ]
    reserved_rows = db.query(
        ReservationItem.item_model_id,
        func.sum(ReservationItem.quantity_requested).label("total_reserved"),
    ).join(Reservation).filter(
        Reservation.date == target_date,
        Reservation.status.in_(active_statuses),
    ).group_by(ReservationItem.item_model_id).all()

    reserved_map = {row.item_model_id: row.total_reserved for row in reserved_rows}
    models = db.query(ItemModel).filter(ItemModel.deleted_at == None).all()
    return [
        {
            "id": m.id, "name": m.name, "category": m.category,
            "description": m.description, "image_url": m.image_url,
            "total_stock": m.total_stock,
            "available_qty": max(0, m.total_stock - reserved_map.get(m.id, 0) - m.maintenance_stock),
        }
        for m in models
    ]


@router.post("/item-models", status_code=status.HTTP_201_CREATED)
async def create_item_model(
    payload: ItemModelCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_MANAGE_ITEMS)), # <-- Trava Adicionada
):
    cat_val = payload.category.value if hasattr(payload.category, "value") else payload.category
    model = ItemModel(
        name=payload.name, category=cat_val, description=payload.description,
        model_number=payload.model_number or None,
        image_url=payload.image_url, total_stock=payload.total_stock,
    )
    db.add(model)
    db.commit()
    db.refresh(model)
    await broadcast("INVENTORY_UPDATED", {"id": model.id, "action": "created"})
    return model


@router.patch("/item-models/{model_id}")
async def update_item_model(
    model_id: int,
    payload: ItemModelUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_MANAGE_ITEMS)), # <-- Trava Adicionada
):
    model = db.query(ItemModel).filter(ItemModel.id == model_id, ItemModel.deleted_at == None).first()
    if not model:
        raise HTTPException(status_code=404, detail="Modelo não encontrado.")

    old_data = {
        "name": model.name, "category": model.category,
        "description": model.description, "image_url": model.image_url,
        "total_stock": model.total_stock,
    }

    if payload.name         is not None: model.name         = payload.name
    if payload.category     is not None: model.category     = payload.category.value if hasattr(payload.category, "value") else payload.category
    if payload.description  is not None: model.description  = payload.description
    if payload.model_number is not None: model.model_number = payload.model_number or None
    if payload.image_url    is not None: model.image_url    = payload.image_url
    if payload.total_stock  is not None: model.total_stock  = payload.total_stock

    new_data = {
        "name": model.name, "category": model.category,
        "description": model.description, "image_url": model.image_url,
        "total_stock": model.total_stock,
    }
    audit = AuditLog(
        table_name="item_models",
        record_id=model_id,
        old_data=json.dumps(old_data, ensure_ascii=False),
        new_data=json.dumps(new_data, ensure_ascii=False),
        user_id=current_user.id,
    )
    db.add(audit)
    db.commit()
    await broadcast("INVENTORY_UPDATED", {"id": model_id, "action": "updated"})
    return model


@router.get("/stock")
async def get_realtime_stock(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    in_use_rows = db.query(
        ReservationItem.item_model_id,
        func.sum(ReservationItem.quantity_requested).label("total"),
    ).join(Reservation).filter(
        Reservation.status == ReservationStatus.EM_USO.value,
    ).group_by(ReservationItem.item_model_id).all()
    in_use_map = {r.item_model_id: r.total for r in in_use_rows}

    loan_rows = db.query(
        InstitutionLoan.item_model_id,
        func.sum(InstitutionLoan.quantity_delivered - InstitutionLoan.quantity_returned).label("total"),
    ).filter(InstitutionLoan.status == "em_aberto").group_by(InstitutionLoan.item_model_id).all()
    loan_map = {r.item_model_id: r.total for r in loan_rows}

    models = db.query(ItemModel).filter(ItemModel.deleted_at == None).all()
    return [
        {
            "id": m.id, "name": m.name, "category": m.category,
            "description": m.description, "image_url": m.image_url,
            "total_stock": m.total_stock,
            "in_use": in_use_map.get(m.id, 0),
            "in_loans": loan_map.get(m.id, 0),
            "maintenance_stock": m.maintenance_stock,
            "available_qty": max(0, m.total_stock - in_use_map.get(m.id, 0) - loan_map.get(m.id, 0) - m.maintenance_stock),
        }
        for m in models
    ]


@router.get("/movements")
async def list_inventory_movements(
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_ALMOX)),
):
    movements = db.query(InventoryMovement).options(
        joinedload(InventoryMovement.model),
        joinedload(InventoryMovement.operator),
    ).order_by(InventoryMovement.created_at.desc()).limit(1000).all()
    return [
        {
            "id": m.id,
            "item_model_id": m.item_model_id,
            "action": m.action,
            "quantity": m.quantity,
            "operator_id": m.operator_id,
            "target": m.target,
            "reservation_id": m.reservation_id,
            "loan_id": m.loan_id,
            "observation": m.observation,
            "created_at": m.created_at.isoformat(),
            "model": {"id": m.model.id, "name": m.model.name, "category": m.model.category} if m.model else None,
            "operator": {"id": m.operator.id, "full_name": m.operator.full_name, "role": m.operator.role.value if hasattr(m.operator.role, 'value') else str(m.operator.role)} if m.operator else None,
        }
        for m in movements
    ]


@router.get("/pending-requests")
async def list_pending_material_requests(
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_ALMOX)),
):
    approved_statuses = [ReservationStatus.APROVADO.value, ReservationStatus.APROVADO_COM_RESSALVAS.value]
    reservations = db.query(Reservation).options(
        joinedload(Reservation.items).joinedload(ReservationItem.model),
        joinedload(Reservation.user),
        joinedload(Reservation.laboratory),
        joinedload(Reservation.slots),
    ).filter(
        Reservation.status.in_(approved_statuses),
        Reservation.items.any(),
    ).order_by(Reservation.date).all()
    return [
        {
            "id": r.id,
            "date": r.date.isoformat() if r.date else None,
            "status": r.status,
            "group_id": r.group_id,
            "user": {"id": r.user.id, "full_name": r.user.full_name, "role": r.user.role} if r.user else None,
            "laboratory": {"id": r.laboratory.id, "name": r.laboratory.name, "block": r.laboratory.block} if r.laboratory else None,
            "slots": [{"id": s.id, "code": s.code, "start_time": s.start_time, "end_time": s.end_time} for s in r.slots],
            "items": [
                {
                    "id": i.id,
                    "item_model_id": i.item_model_id,
                    "physical_item_id": i.physical_item_id,
                    "quantity_requested": i.quantity_requested,
                    "quantity_returned": i.quantity_returned,
                    "return_status": i.return_status,
                    "damage_observation": i.damage_observation,
                    "model": {"id": i.model.id, "name": i.model.name, "category": i.model.category} if i.model else None,
                }
                for i in r.items
            ],
        }
        for r in reservations
    ]


class MaintenanceResolvePayload(BaseModel):
    qty_repaired: int = 0
    qty_discarded: int = 0
    observation: Optional[str] = None

@router.post("/item-models/{model_id}/resolve-maintenance", tags=["inventário"])
async def resolve_maintenance(
    model_id: int,
    payload: MaintenanceResolvePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_MANAGE_ITEMS)), # <-- Trava Adicionada
):
    model = db.query(ItemModel).filter(ItemModel.id == model_id, ItemModel.deleted_at == None).first()
    if not model:
        raise HTTPException(status_code=404, detail="Item não encontrado.")

    total_resolved = payload.qty_repaired + payload.qty_discarded
    if total_resolved <= 0:
        raise HTTPException(status_code=400, detail="Informe uma quantidade a ser reparada ou descartada.")
    if total_resolved > model.maintenance_stock:
        raise HTTPException(status_code=400, detail="Quantidade informada excede o estoque em manutenção.")

    model.maintenance_stock -= total_resolved
    model.total_stock = max(0, model.total_stock - payload.qty_discarded)

    if payload.qty_repaired > 0:
        db.add(InventoryMovement(
            item_model_id=model_id,
            action="reparo",
            quantity=payload.qty_repaired,
            operator_id=current_user.id,
            target="Estoque Disponível",
            observation=payload.observation
        ))
    if payload.qty_discarded > 0:
        db.add(InventoryMovement(
            item_model_id=model_id,
            action="descarte",
            quantity=payload.qty_discarded,
            operator_id=current_user.id,
            target="Descarte / Lixo",
            observation=payload.observation
        ))

    db.commit()
    await broadcast("INVENTORY_UPDATED", {"id": model_id, "action": "maintenance_resolved"})
    return {"message": "Manutenção resolvida com sucesso."}


CATEGORY_MAP = {
    "eletronica": "eletronica",
    "eletrônica": "eletronica",
    "fisica": "fisica",
    "física": "fisica",
    "automacao": "automacao",
    "automação": "automacao",
    "eletrica": "eletrica",
    "elétrica": "eletrica",
    "componentes": "componentes",
}


class ImportConfirmPayload(BaseModel):
    items: list


@router.post("/import/preview")
async def import_inventory_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleChecker(_MANAGE_ITEMS)),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Instale openpyxl: pip install openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header_row_idx = None
    col_map = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "categoria" in row_lower:
            header_row_idx = row_idx
            for ci, val in enumerate(row_lower):
                if val in ("equipamentos", "nome", "item", "descrição", "name"):
                    col_map["name"] = ci
                elif val == "categoria":
                    col_map["category"] = ci
                elif val in ("modelo", "model", "referência", "ref"):
                    col_map["model_number"] = ci
                elif val in ("quantidade", "qtd", "qty", "qtde"):
                    col_map["total_stock"] = ci
            break

    if header_row_idx is None or "category" not in col_map:
        raise HTTPException(status_code=400, detail="Planilha inválida. Necessário coluna 'Categoria'.")

    if "name" not in col_map:
        col_map["name"] = 0

    items = []
    errors = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue

        name_raw = row[col_map["name"]] if len(row) > col_map["name"] else None
        if not name_raw or str(name_raw).strip() == "":
            continue

        name = str(name_raw).strip()

        cat_raw = ""
        if "category" in col_map and len(row) > col_map["category"] and row[col_map["category"]]:
            cat_raw = str(row[col_map["category"]]).strip()

        category = CATEGORY_MAP.get(cat_raw.lower())
        if not category:
            errors.append(f"Categoria '{cat_raw}' desconhecida para '{name}' — item ignorado.")
            continue

        model_number = ""
        if "model_number" in col_map and len(row) > col_map["model_number"] and row[col_map["model_number"]]:
            model_number = str(row[col_map["model_number"]]).strip()

        total_stock = 0
        if "total_stock" in col_map and len(row) > col_map["total_stock"] and row[col_map["total_stock"]]:
            try:
                total_stock = int(float(str(row[col_map["total_stock"]]).strip()))
            except (ValueError, TypeError):
                total_stock = 0

        items.append({
            "name": name,
            "category": category,
            "model_number": model_number,
            "total_stock": total_stock,
        })

    return {"items": items, "errors": errors, "total": len(items)}


@router.post("/import/confirm", status_code=status.HTTP_201_CREATED)
async def import_inventory_confirm(
    payload: ImportConfirmPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(_MANAGE_ITEMS)),
):
    created = 0
    skipped = 0
    for item in payload.items:
        if not item.get("name") or not item.get("category"):
            skipped += 1
            continue
        existing = db.query(ItemModel).filter(
            ItemModel.name == item["name"],
            ItemModel.deleted_at == None
        ).first()
        if existing:
            skipped += 1
            continue
        cat = CATEGORY_MAP.get(str(item["category"]).lower(), item["category"])
        new_model = ItemModel(
            name=item["name"],
            category=cat,
            model_number=item.get("model_number") or None,
            total_stock=int(item.get("total_stock", 0)),
        )
        db.add(new_model)
        created += 1
    db.commit()
    await broadcast("INVENTORY_UPDATED", {"action": "bulk_import", "count": created})
    return {"created": created, "skipped": skipped}