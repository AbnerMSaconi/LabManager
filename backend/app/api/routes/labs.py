import json
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import List, Optional
from pydantic import BaseModel

from ..deps import get_db, RoleChecker, get_current_user
from ...models.base_models import (
    Laboratory, Software as SoftwareModel, LessonSlot,
    Reservation, ReservationSlot, ReservationStatus, User, UserRole, AuditLog
)
from ...schemas.reservation_schemas import LaboratoryCreate, LaboratoryUpdate, SoftwareCreate

router = APIRouter(prefix="/api/v1", tags=["laboratórios"])


class LabAvailabilityRequest(BaseModel):
    dates: List[str]
    slot_ids: List[int]
    block: Optional[str] = None


@router.post("/labs/available")
async def get_available_labs(
    req: LabAvailabilityRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import datetime as _dt
    if not req.dates or not req.slot_ids:
        raise HTTPException(status_code=400, detail="Datas e horários são obrigatórios.")

    parsed_dates = []
    for d in req.dates:
        try:
            parsed_dates.append(_dt.strptime(d, "%Y-%m-%d").date())
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Data inválida: {d}")

    active_statuses = [
        ReservationStatus.PENDENTE.value,
        ReservationStatus.APROVADO.value,
        ReservationStatus.AGUARDANDO_SOFTWARE.value,
        ReservationStatus.EM_USO.value,
    ]

    conflicting_rows = (
        db.query(Reservation.lab_id)
        .join(ReservationSlot, ReservationSlot.reservation_id == Reservation.id)
        .filter(
            Reservation.date.in_(parsed_dates),
            Reservation.status.in_(active_statuses),
            ReservationSlot.slot_id.in_(req.slot_ids),
        )
        .distinct()
        .all()
    )
    conflicting_ids = {row.lab_id for row in conflicting_rows}

    query = db.query(Laboratory).options(joinedload(Laboratory.softwares))
    if req.block:
        query = query.filter(Laboratory.block == req.block)
    if conflicting_ids:
        query = query.filter(Laboratory.id.notin_(conflicting_ids))

    return query.order_by(Laboratory.name).all()


@router.get("/labs")
async def list_labs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(Laboratory).options(
        joinedload(Laboratory.softwares)
    ).filter(Laboratory.deleted_at == None).all()


@router.get("/labs/{lab_id}")
async def get_lab(
    lab_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    lab = db.query(Laboratory).options(
        joinedload(Laboratory.softwares)
    ).filter(Laboratory.id == lab_id, Laboratory.deleted_at == None).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratório não encontrado.")
    return lab


@router.post("/labs", status_code=status.HTTP_201_CREATED)
async def create_lab(
    payload: LaboratoryCreate,
    db: Session = Depends(get_db),
    # PROGEX Removido
    current_user: User = Depends(RoleChecker([UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]))
):
    VALID_BLOCKS = {"Bloco A", "Bloco B", "Bloco C"}
    if payload.block not in VALID_BLOCKS:
        raise HTTPException(status_code=400, detail=f"Bloco inválido: {payload.block}. Use: {', '.join(sorted(VALID_BLOCKS))}")
    lab = Laboratory(
        name=payload.name, block=payload.block, room_number=payload.room_number,
        capacity=payload.capacity, is_practical=payload.is_practical,
        description=payload.description,
    )
    db.add(lab)
    db.flush()
    for sw_id in payload.software_ids:
        sw = db.query(SoftwareModel).filter(SoftwareModel.id == sw_id).first()
        if sw:
            lab.softwares.append(sw)
    db.commit()
    db.refresh(lab)
    return {"message": "Laboratório criado.", "id": lab.id}


@router.put("/labs/{lab_id}")
async def update_lab(
    lab_id: int,
    payload: LaboratoryUpdate,
    db: Session = Depends(get_db),
    # PROGEX Removido
    current_user: User = Depends(RoleChecker([UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]))
):
    lab = db.query(Laboratory).options(joinedload(Laboratory.softwares)).filter(Laboratory.id == lab_id, Laboratory.deleted_at == None).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratório não encontrado.")

    old_data = {
        "name": lab.name, "block": lab.block, "room_number": lab.room_number,
        "capacity": lab.capacity, "is_practical": lab.is_practical,
        "description": lab.description, "is_active": lab.is_active
    }

    VALID_BLOCKS = {"Bloco A", "Bloco B", "Bloco C"}
    if payload.name        is not None: lab.name        = payload.name
    if payload.block       is not None:
        if payload.block not in VALID_BLOCKS:
            raise HTTPException(status_code=400, detail=f"Bloco inválido: {payload.block}. Use: {', '.join(sorted(VALID_BLOCKS))}")
        lab.block = payload.block
    if payload.room_number  is not None: lab.room_number  = payload.room_number
    if payload.capacity     is not None: lab.capacity     = payload.capacity
    if payload.is_practical is not None: lab.is_practical = payload.is_practical
    if payload.description  is not None: lab.description  = payload.description
    if payload.is_active    is not None: lab.is_active    = payload.is_active
    if payload.software_ids is not None:
        lab.softwares = []
        for sw_id in payload.software_ids:
            sw = db.query(SoftwareModel).filter(SoftwareModel.id == sw_id).first()
            if sw: lab.softwares.append(sw)

    new_data = {
        "name": lab.name, "block": lab.block, "room_number": lab.room_number,
        "capacity": lab.capacity, "is_practical": lab.is_practical,
        "description": lab.description, "is_active": lab.is_active
    }
    audit = AuditLog(
        table_name="laboratories",
        record_id=lab_id,
        old_data=json.dumps(old_data, ensure_ascii=False),
        new_data=json.dumps(new_data, ensure_ascii=False),
        user_id=current_user.id,
    )
    db.add(audit)
    db.commit()
    return {"message": "Laboratório atualizado."}


@router.delete("/labs/{lab_id}")
async def delete_lab(
    lab_id: int,
    db: Session = Depends(get_db),
    # PROGEX Removido
    current_user: User = Depends(RoleChecker([UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]))
):
    lab = db.query(Laboratory).filter(Laboratory.id == lab_id, Laboratory.deleted_at == None).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratório não encontrado.")
    lab.deleted_at = func.now()
    db.commit()
    return {"message": "Laboratório movido para a quarentena."}


@router.get("/slots")
async def list_slots(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(LessonSlot).order_by(LessonSlot.code).all()


@router.get("/softwares")
async def list_softwares(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(SoftwareModel).filter(SoftwareModel.deleted_at == None).all()


@router.post("/softwares", status_code=status.HTTP_201_CREATED)
async def create_software(
    payload: SoftwareCreate,
    db: Session = Depends(get_db),
    # PROGEX Removido
    current_user: User = Depends(RoleChecker([UserRole.DTI_TECNICO, UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]))
):
    sw = SoftwareModel(name=payload.name, version=payload.version)
    db.add(sw)
    db.commit()
    db.refresh(sw)
    return {"message": "Software cadastrado.", "id": sw.id}


@router.delete("/softwares/{sw_id}")
async def delete_software(
    sw_id: int,
    db: Session = Depends(get_db),
    # PROGEX Removido
    current_user: User = Depends(RoleChecker([UserRole.DTI_TECNICO, UserRole.ADMINISTRADOR, UserRole.SUPER_ADMIN]))
):
    sw = db.query(SoftwareModel).filter(SoftwareModel.id == sw_id, SoftwareModel.deleted_at == None).first()
    if not sw:
        raise HTTPException(status_code=404, detail="Software não encontrado.")
    sw.deleted_at = func.now()
    db.commit()
    return {"message": "Software movido para a quarentena."}


class SoftwareImportConfirmItem(BaseModel):
    lab_id: Optional[int] = None
    lab_name: str
    softwares: List[str]


class SoftwareImportConfirmPayload(BaseModel):
    items: List[SoftwareImportConfirmItem]


@router.post("/softwares/import/preview")
async def import_softwares_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Instale openpyxl: pip install openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    result = []

    # Format 1: tab-per-lab (each sheet name = lab, column "Software" or "Nome" = software names)
    if len(wb.sheetnames) > 1 or (len(wb.sheetnames) == 1 and wb.sheetnames[0].lower() not in ("sheet1", "sheet", "planilha1", "planilha")):
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lab_name = sheet_name.strip()
            # Find software column
            sw_col_idx = None
            header_row_idx = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
                if any(v in ("software", "nome", "name", "softwares") for v in row_lower):
                    header_row_idx = row_idx
                    for ci, v in enumerate(row_lower):
                        if v in ("software", "nome", "name", "softwares"):
                            sw_col_idx = ci
                            break
                    break
            softwares = []
            if header_row_idx is not None and sw_col_idx is not None:
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if row and len(row) > sw_col_idx and row[sw_col_idx]:
                        sw = str(row[sw_col_idx]).strip()
                        if sw and sw not in softwares:
                            softwares.append(sw)
            elif header_row_idx is None:
                # No header — assume first column is software names
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if row and row[0]:
                        sw = str(row[0]).strip()
                        if sw and sw not in softwares:
                            softwares.append(sw)
            if softwares:
                result.append({"lab_name": lab_name, "softwares": softwares, "lab_id": None})
    else:
        # Format 2: single sheet with lab names as column headers
        ws = wb.active
        header_row_idx = None
        lab_columns = {}  # col_index -> lab_name

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            has_lab = any(
                c and ("laboratório" in str(c).lower() or "lab" in str(c).lower())
                for c in row if c
            )
            if has_lab:
                header_row_idx = row_idx
                for ci, cell in enumerate(row):
                    if cell and str(cell).strip():
                        lab_columns[ci] = str(cell).strip()
                break

        if not header_row_idx:
            raise HTTPException(status_code=400, detail="Planilha inválida. Use abas por laboratório ou cabeçalho com nomes dos labs.")

        lab_softwares = {name: [] for name in lab_columns.values()}

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            for ci, lab_name in lab_columns.items():
                if len(row) > ci and row[ci] and str(row[ci]).strip():
                    sw_name = str(row[ci]).strip()
                    if sw_name not in lab_softwares[lab_name]:
                        lab_softwares[lab_name].append(sw_name)

        result = [
            {"lab_name": lab_name, "softwares": softwares, "lab_id": None}
            for lab_name, softwares in lab_softwares.items()
            if softwares
        ]

    return {"labs": result, "total_softwares": sum(len(x["softwares"]) for x in result)}


@router.post("/softwares/import/confirm")
async def import_softwares_confirm(
    payload: SoftwareImportConfirmPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    created_sw = 0
    linked = 0

    for item in payload.items:
        lab = None
        if item.lab_id:
            lab = db.query(Laboratory).filter(Laboratory.id == item.lab_id).first()

        for sw_name in item.softwares:
            if not sw_name.strip():
                continue
            existing = db.query(SoftwareModel).filter(SoftwareModel.name == sw_name).first()
            if not existing:
                existing = SoftwareModel(name=sw_name)
                db.add(existing)
                db.flush()
                created_sw += 1
            if lab and existing not in lab.softwares:
                lab.softwares.append(existing)
                linked += 1

    db.commit()
    return {"created": created_sw, "linked": linked}